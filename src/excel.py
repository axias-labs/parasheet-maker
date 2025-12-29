from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter


# ============================================================
# Markdown → Excel 変換ロジック
# ============================================================

def sanitize_sheet_name(title: str) -> str:
    """Excelのシート名として使えるように整形"""
    invalid = '[]:*?/\\'
    name = "".join(c for c in title if c not in invalid).strip()
    if not name:
        name = "Sheet1"
    if len(name) > 31:
        name = name[:31]
    return name


def is_separator_row(cells) -> bool:
    """Markdownテーブルの区切り行（---など）かどうか判定"""
    for c in cells:
        s = c.strip().replace(":", "").replace(" ", "")
        if not s:
            return False
        if any(ch != "-" for ch in s):
            return False
    return True


def parse_markdown_table(table_lines):
    """Markdownのテーブル部分（|...| の連続）を2次元配列に変換"""
    rows = []
    for line in table_lines:
        s = line.strip()
        if not s.startswith("|"):
            continue
        if s.endswith("|"):
            s = s[1:-1]
        else:
            s = s[1:]
        cells = [c.strip() for c in s.split("|")]
        if is_separator_row(cells):
            continue
        rows.append(cells)
    return rows


def markdown_to_sheets(md_text: str):
    """
    Markdownテキストから
    {シート名: [ [セル1,セル2,...], ... ] } の形に変換
    """
    sheets = {}
    lines = md_text.splitlines()
    current_sheet = "Sheet1"

    i = 0
    while i < len(lines):
        line = lines[i]

        # 見出し行（# で始まる）をシート名として使う
        stripped = line.lstrip("#").strip()
        if line.startswith("#") and stripped:
            current_sheet = sanitize_sheet_name(stripped)
            sheets.setdefault(current_sheet, [])
            i += 1
            continue

        # resource_type見出し（独自マーカー）
        if line.strip().startswith("@resource_type "):
            rt = line.strip().split(" ", 1)[1].strip()
            sheet_rows = sheets.setdefault(current_sheet, [])
            # 表の前にスペースが欲しければ空行
            if sheet_rows and sheet_rows[-1] != []:
                sheet_rows.append([])
            sheet_rows.append([rt])  # 1セル行
            i += 1
            continue

        # テーブル行を検出
        if line.strip().startswith("|") and "|" in line.strip()[1:]:
            # テーブルブロックをまとめて取得
            table_block = [line]
            j = i + 1
            while j < len(lines) and lines[j].strip().startswith("|"):
                table_block.append(lines[j])
                j += 1

            rows = parse_markdown_table(table_block)
            if rows:
                sheet_rows = sheets.setdefault(current_sheet, [])
                if sheet_rows:
                    # 複数テーブルが同じシートに来たときの区切り用に1行空ける
                    sheet_rows.append([])
                sheet_rows.extend(rows)

            # テーブル全体を処理し終わったので、iをまとめて進める
            i = j
            continue

        # テーブルでも見出しでもないので次の行へ
        i += 1

    return sheets


def try_format_json_object_list(text: str):
    """
    text が「JSONとして有効な、オブジェクト(dict)の配列」([{}, {} ...]) の場合だけ、
    インデント付きの複数行JSONに整形して返す。
    それ以外（JSONでない / dict配列でない）は None を返す。
    """
    s = str(text).strip()
    if not (s.startswith("[") and s.endswith("]")):
        return None

    try:
        parsed = json.loads(s)
    except Exception:
        return None

    if not isinstance(parsed, list):
        return None

    if not parsed or not all(isinstance(x, dict) for x in parsed):
        return None

    pretty = json.dumps(parsed, ensure_ascii=False, indent=2)
    return pretty


def write_excel_from_markdown(md_text: str, excel_path: Path):
    sheets = markdown_to_sheets(md_text)

    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    header_fill = PatternFill(fill_type="solid", start_color="CCFFCC", end_color="CCFFCC")

    if not sheets:
        ws = wb.create_sheet("Sheet1")
        ws["A1"] = "（Markdownからテーブルを検出できませんでした）"
        ws["A1"].border = border
        ws["A1"].fill = header_fill
    else:
        for idx, (sheet_name, rows) in enumerate(sheets.items()):
            ws = wb.create_sheet(sheet_name if sheet_name else f"Sheet{idx+1}")

            header_next = True

            for r_idx, row in enumerate(rows, start=1):

                # 空行 → 次の非空行はヘッダ
                if row == []:
                    header_next = True
                    continue

                # resource_type見出し行（1セル行）
                is_rt_title = (len(row) == 1 and row[0] is not None and str(row[0]).strip() != "")

                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)

                    # resource_type見出し行（タイトル）
                    if is_rt_title:
                        cell.font = Font(name=cell.font.name, size=14, bold=True)
                        cell.alignment = Alignment(vertical="top", wrap_text=False)
                        cell.border = Border()  # 罫線なし
                        continue

                    # ここから下は「通常セル」
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

                    formatted = try_format_json_object_list(cell.value)
                    if formatted is not None:
                        cell.value = formatted
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

                    cell.border = border

                    # 表ヘッダのスタイル（各表の先頭行）
                    if header_next:
                        cell.fill = PatternFill(
                            fill_type="solid",
                            start_color="006400",
                            end_color="006400",
                        )
                        cell.font = Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=True,
                            color="FFFFFFFF",
                        )

                # 行処理後のフラグ更新
                if is_rt_title:
                    header_next = True
                else:
                    header_next = False

    # 列幅を自動調整
    for ws in wb.worksheets:
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)

            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    value_str = str(cell.value)

                    if "\n" in value_str:
                        line_lengths = [len(line) for line in value_str.splitlines()]
                        length = max(line_lengths) if line_lengths else 0
                    else:
                        length = len(value_str)

                    if length > max_length:
                        max_length = length

            if max_length > 0:
                ws.column_dimensions[col_letter].width = max_length + 2

    # 全シートの枠線 = Gridlines を非表示
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    wb.save(excel_path)
